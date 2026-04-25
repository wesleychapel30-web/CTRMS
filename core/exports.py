"""
Export views for PDF, Excel, and CSV formats
Handles data export for reports and requests
"""
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from requests.models import Request
from invitations.models import Invitation
from core.rbac import require_permission, user_has_permission
import csv
import json
from datetime import datetime
from io import BytesIO
from django.utils.dateparse import parse_datetime, parse_date
from django.db.models import Q

# Try to import Excel and PDF libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def _format_tzs(value):
    amount = float(value or 0)
    return f"TZS {amount:,.2f}"


def _filtered_requests_queryset(request):
    """Apply the same basic filters used in the UI (status/category/search/date range)."""
    status_filter = (request.GET.get('status') or '').strip()
    category_filter = (request.GET.get('category') or '').strip()
    search = (request.GET.get('search') or '').strip()

    created_gte_raw = (request.GET.get('created_at__gte') or '').strip()
    created_lte_raw = (request.GET.get('created_at__lte') or '').strip()

    queryset = Request.objects.all().order_by('-created_at')

    # Enforce data visibility rules on exports.
    user = getattr(request, "user", None)
    if user and user.is_authenticated and not user_has_permission(user, "request:view_all"):
        if user_has_permission(user, "payment:view") or user_has_permission(user, "payment:record"):
            queryset = queryset.filter(status__in=[
                Request.Status.DIRECTOR_APPROVED,
                Request.Status.FINANCE_PROCESSING,
                Request.Status.FINANCE_QUERY,
                Request.Status.PENDING_PAYMENT,
                Request.Status.APPROVED,
                Request.Status.PARTIALLY_PAID,
                Request.Status.PAID,
            ])
        else:
            queryset = queryset.filter(created_by=user)

    if status_filter:
        queryset = queryset.filter(status=status_filter)
    if category_filter:
        queryset = queryset.filter(category=category_filter)
    if search:
        queryset = queryset.filter(
            Q(request_id__icontains=search)
            | Q(applicant_name__icontains=search)
            | Q(applicant_email__icontains=search)
        )

    if created_gte_raw:
        created_gte = parse_datetime(created_gte_raw) or (
            datetime.combine(parse_date(created_gte_raw), datetime.min.time()) if parse_date(created_gte_raw) else None
        )
        if created_gte is not None:
            queryset = queryset.filter(created_at__gte=created_gte)
    if created_lte_raw:
        created_lte = parse_datetime(created_lte_raw) or (
            datetime.combine(parse_date(created_lte_raw), datetime.max.time()) if parse_date(created_lte_raw) else None
        )
        if created_lte is not None:
            queryset = queryset.filter(created_at__lte=created_lte)

    return queryset


@login_required
@require_http_methods(["GET"])
@require_permission("report:export")
def export_requests_csv(request):
    """
    Export all requests to CSV format
    """
    try:
        queryset = _filtered_requests_queryset(request)
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="requests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Request ID', 'Applicant', 'Phone', 'Category', 'Amount Requested',
            'Amount Approved', 'Amount Disbursed', 'Status', 'Created Date', 'Approved Date'
        ])
        
        for req in queryset:
            writer.writerow([
                req.request_id,
                req.applicant_name,
                req.applicant_phone,
                req.category,
                float(req.amount_requested),
                float(req.approved_amount or 0),
                float(req.disbursed_amount or 0),
                req.status,
                req.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                req.reviewed_at.strftime('%Y-%m-%d %H:%M:%S') if req.reviewed_at else ''
            ])
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
@require_permission("report:export")
def export_requests_excel(request):
    """
    Export all requests to Excel format
    """
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({
            'success': False,
            'error': 'Excel export not available. Install openpyxl package.'
        }, status=400)
    
    try:
        queryset = _filtered_requests_queryset(request)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Requests"
        
        # Style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = [
            'Request ID', 'Applicant', 'Phone', 'Category', 'Amount Requested',
            'Amount Approved', 'Amount Disbursed', 'Status', 'Created Date', 'Approved Date'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_idx, req in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1).value = req.request_id
            ws.cell(row=row_idx, column=2).value = req.applicant_name
            ws.cell(row=row_idx, column=3).value = req.applicant_phone
            ws.cell(row=row_idx, column=4).value = req.category
            ws.cell(row=row_idx, column=5).value = float(req.amount_requested)
            ws.cell(row=row_idx, column=6).value = float(req.approved_amount or 0)
            ws.cell(row=row_idx, column=7).value = float(req.disbursed_amount or 0)
            ws.cell(row=row_idx, column=8).value = req.status
            ws.cell(row=row_idx, column=9).value = req.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_idx, column=10).value = req.reviewed_at.strftime('%Y-%m-%d %H:%M:%S') if req.reviewed_at else ''
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 18
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="requests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        response.write(output.getvalue())
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
@require_permission("report:export")
def export_requests_pdf(request):
    """
    Export all requests to PDF format
    """
    if not REPORTLAB_AVAILABLE:
        return JsonResponse({
            'success': False,
            'error': 'PDF export not available. Install reportlab package.'
        }, status=400)
    
    try:
        queryset = _filtered_requests_queryset(request)
        
        # Create PDF
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=20
        )
        
        # Title
        story.append(Paragraph("Request Report", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Summary
        summary_text = f"""
        <b>Report Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
        <b>Total Requests:</b> {queryset.count()}<br/>
        <b>Total Amount Requested:</b> {_format_tzs(queryset.aggregate(Sum('amount_requested'))['amount_requested__sum'])}<br/>
        <b>Total Amount Approved:</b> {_format_tzs(queryset.aggregate(Sum('approved_amount'))['approved_amount__sum'])}
        """
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Table data
        table_data = [['Request ID', 'Applicant', 'Category', 'Requested', 'Approved', 'Status']]
        
        for req in queryset[:50]:  # Limit to 50 rows for readability
            table_data.append([
                req.request_id,
                req.applicant_name[:20],
                req.category,
                _format_tzs(req.amount_requested),
                _format_tzs(req.approved_amount or 0),
                req.status
            ])
        
        # Create table
        table = Table(table_data, colWidths=[1.2*inch, 1.8*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="requests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        response.write(output.getvalue())
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
@require_permission("report:export")
def export_financial_report_csv(request):
    """
    Export financial report to CSV
    """
    try:
        queryset = _filtered_requests_queryset(request).filter(
            status__in=[Request.Status.APPROVED, Request.Status.PARTIALLY_PAID, Request.Status.PAID]
        )

        # Get financial data
        requests_data = queryset.values('category').annotate(
            count=Count('id'),
            total_amount=Sum('approved_amount')
        )
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="financial_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Category', 'Count', 'Total Amount'])
        
        for row in requests_data:
            writer.writerow([
                row['category'],
                row['count'],
                float(row['total_amount'] or 0)
            ])
        
        # Summary row
        totals = queryset.aggregate(
            count=Count('id'),
            total_amount=Sum('approved_amount')
        )
        writer.writerow(['TOTAL', totals['count'], float(totals['total_amount'] or 0)])
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
@require_permission("report:export")
def export_analytics_report_csv(request):
    """
    Export analytics report to CSV
    """
    try:
        queryset = _filtered_requests_queryset(request)
        total_requests = queryset.count()
        draft_requests = queryset.filter(status=Request.Status.DRAFT).count()
        approved_requests = queryset.filter(status=Request.Status.APPROVED).count()
        rejected_requests = queryset.filter(status=Request.Status.REJECTED).count()
        pending_requests = queryset.filter(status=Request.Status.PENDING).count()
        partially_paid_requests = queryset.filter(status=Request.Status.PARTIALLY_PAID).count()
        paid_requests = queryset.filter(status=Request.Status.PAID).count()
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="analytics_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Requests', total_requests])
        writer.writerow(['Draft Requests', draft_requests])
        writer.writerow(['Approved Requests', approved_requests])
        writer.writerow(['Rejected Requests', rejected_requests])
        writer.writerow(['Pending Requests', pending_requests])
        writer.writerow(['Partially Paid Requests', partially_paid_requests])
        writer.writerow(['Paid Requests', paid_requests])
        writer.writerow(['Approval Rate (%)', round((approved_requests/total_requests*100), 2) if total_requests > 0 else 0])
        writer.writerow(['Generated Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
